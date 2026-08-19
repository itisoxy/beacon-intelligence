from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import duckdb
import pandas as pd
from openpyxl import load_workbook

from .business_tools import tool_schemas
from .metrics import METRIC_PERIODS, build_metric_values, metric_registry_records
from .research import (
    HORIZONS,
    build_research_by_horizon,
    fund_horizon_view,
    allocation_horizon_view,
    manager_horizon_view,
    q4_vs_q3,
    h2_vs_h1,
    q1_to_q4_trajectory,
)
from .semantic import build_semantic_layer


STRUCTURED_SHEETS = {
    "Fund_Summary",
    "Asset_Allocation",
    "Manager_Detail",
    "Cash_Flow_Detail",
    "Benchmarks_Reference",
}
PERIODS = ["Q1", "Q2", "Q3", "Q4"]
ROUNDING_TOLERANCE = 0.05


@dataclass(frozen=True)
class SourceWorkbook:
    path: Path

    @property
    def name(self) -> str:
        return self.path.name

    @property
    def as_of(self) -> str:
        stem = self.path.stem[:8]
        return f"{stem[:4]}-{stem[4:6]}-{stem[6:8]}"


def _json_value(value: Any) -> Any:
    if isinstance(value, list):
        return [_json_value(item) for item in value]
    if isinstance(value, dict):
        return {str(key): _json_value(item) for key, item in value.items()}
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.date().isoformat()
    if hasattr(value, "item"):
        return value.item()
    return value


def _records(df: pd.DataFrame) -> list[dict[str, Any]]:
    clean = df.where(pd.notna(df), None)
    return [{key: _json_value(value) for key, value in row.items()} for row in clean.to_dict(orient="records")]


def _prov(row: pd.Series) -> dict[str, Any]:
    value = row.get("_provenance")
    return value if isinstance(value, dict) else {}


def _base_lineage(row: pd.Series, record_id: str, fiscal_year: str = "FY2026") -> dict[str, Any]:
    provenance = _prov(row)
    return {
        "record_id": record_id,
        "source_record_id": row.get("source_record_id"),
        "source_file": provenance.get("source_file"),
        "source_sheet": provenance.get("source_sheet"),
        "source_row": provenance.get("source_row"),
        "source_cells": provenance.get("source_cells"),
        "fiscal_year": fiscal_year,
        "quarter": row.get("Quarter"),
        "reporting_period": row.get("Horizon") or row.get("Quarter"),
    }


def _source_fields(row: pd.Series, fields: list[str]) -> dict[str, Any]:
    return {field: _json_value(row.get(field)) for field in fields if field in row}


def _cell_range(row_number: int, column_count: int) -> str:
    def col_name(index: int) -> str:
        result = ""
        while index:
            index, rem = divmod(index - 1, 26)
            result = chr(65 + rem) + result
        return result

    return f"A{row_number}:{col_name(column_count)}{row_number}"


def read_readme(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "ReadMe" not in wb.sheetnames:
            return []
        ws = wb["ReadMe"]
        lines: list[str] = []
        for row in ws.iter_rows(values_only=True):
            text = " ".join(str(cell) for cell in row if cell not in (None, ""))
            if text:
                lines.append(text)
        return lines
    finally:
        wb.close()


def read_workbook(source: SourceWorkbook) -> tuple[dict[str, pd.DataFrame], dict[str, Any]]:
    readme = read_readme(source.path)
    xls = pd.ExcelFile(source.path, engine="openpyxl")
    sheet_meta: list[dict[str, Any]] = []
    frames: dict[str, pd.DataFrame] = {}
    for sheet in xls.sheet_names:
        if sheet == "ReadMe":
            sheet_meta.append({"name": sheet, "rows": len(readme), "cols": 1, "header": []})
            continue
        if sheet == "RAW_Export_Extract":
            raw = pd.read_excel(source.path, sheet_name=sheet, header=None, engine="openpyxl")
            sheet_meta.append({"name": sheet, "rows": int(len(raw)), "cols": int(raw.shape[1]), "header": []})
            frames[sheet] = raw
            continue
        df = pd.read_excel(source.path, sheet_name=sheet, engine="openpyxl")
        for col in df.columns:
            if "Date" in str(col):
                df[col] = pd.to_datetime(df[col]).dt.date.astype(str)
        source_column_count = len(df.columns)
        df["_provenance"] = [
            {
                "source_file": source.name,
                "source_sheet": sheet,
                "source_row": int(idx) + 2,
                "source_cells": _cell_range(int(idx) + 2, source_column_count),
            }
            for idx in range(len(df))
        ]
        frames[sheet] = df
        sheet_meta.append(
            {
                "name": sheet,
                "rows": int(len(df) + 1),
                "cols": int(len(df.columns) - 1),
                "header": [str(col) for col in df.columns if col != "_provenance"],
            }
        )
    meta = {"file": source.name, "as_of": source.as_of, "sheets": sheet_meta, "readme": readme}
    return frames, meta


def source_record_id(row: pd.Series, sheet: str) -> str:
    parts = [sheet, str(row.get("Quarter", "")), str(row.get("QuarterEndDate", "")), str(row.get("FundCode", ""))]
    for field in ("AssetClassLevel1", "ManagerName", "FlowType", "Benchmark"):
        if field in row and pd.notna(row[field]):
            parts.append(str(row[field]))
    return "|".join(parts)


def canonicalize(df: pd.DataFrame, sheet: str) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    if df.empty:
        return df, []
    keyed = df.copy()
    keyed["source_record_id"] = keyed.apply(lambda row: source_record_id(row, sheet), axis=1)
    keyed["_source_file_date"] = keyed["_provenance"].map(lambda p: str(p["source_file"])[:8])
    if "QuarterEndDate" in keyed.columns:
        keyed["_quarter_date"] = keyed["QuarterEndDate"].astype(str).str.replace("-", "", regex=False)
    else:
        keyed["_quarter_date"] = keyed["_source_file_date"]
    duplicates: list[dict[str, Any]] = []
    rows: list[pd.Series] = []
    for key, group in keyed.groupby("source_record_id", sort=False):
        if len(group) > 1:
            duplicates.append(
                {
                    "key": key,
                    "count": int(len(group)),
                    "files": sorted(group["_provenance"].map(lambda p: p["source_file"]).unique().tolist()),
                }
            )
        exact = group[group["_source_file_date"] == group["_quarter_date"]]
        chosen = exact.iloc[0] if not exact.empty else group.sort_values("_source_file_date").iloc[-1]
        rows.append(chosen)
    out = pd.DataFrame(rows).drop(columns=["_source_file_date", "_quarter_date"])
    if "QuarterEndDate" in out.columns:
        out = out.sort_values([col for col in ["QuarterEndDate", "FundCode", "AssetClassLevel1", "ManagerName"] if col in out.columns])
    return out.reset_index(drop=True), duplicates


def load_sources(data_dir: Path) -> tuple[dict[str, pd.DataFrame], list[dict[str, Any]], dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    collected = {sheet: [] for sheet in STRUCTURED_SHEETS}
    books: list[dict[str, Any]] = []
    raw_samples: list[dict[str, Any]] = []
    for path in sorted(data_dir.glob("*.xlsx")):
        frames, meta = read_workbook(SourceWorkbook(path))
        books.append(meta)
        for sheet in STRUCTURED_SHEETS:
            if sheet in frames:
                collected[sheet].append(frames[sheet])
        raw = frames.get("RAW_Export_Extract")
        if raw is not None:
            sample = raw.head(14).where(pd.notna(raw), None).values.tolist()
            raw_samples.append({"source_file": path.name, "sample": sample})

    canonical: dict[str, pd.DataFrame] = {}
    duplicates: dict[str, list[dict[str, Any]]] = {}
    for sheet, frames in collected.items():
        combined = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        canonical[sheet], duplicates[sheet.lower()] = canonicalize(combined, sheet)
    return canonical, books, duplicates, raw_samples


def build_fund_view(fund_summary: pd.DataFrame) -> pd.DataFrame:
    rows = []
    base = fund_summary.copy()
    for _, row in base.iterrows():
        rows.append(row.to_dict())
    for quarter, group in base.groupby("Quarter", sort=False):
        ending = group["EndingMarketValue"].sum()
        weighted = lambda field: float((group[field] * group["EndingMarketValue"]).sum() / ending) if ending else 0.0
        rows.append(
            {
                "Quarter": quarter,
                "QuarterEndDate": group["QuarterEndDate"].iloc[0],
                "FundCode": "All",
                "FundName": "Combined Portfolio",
                "FundType": "Combined reporting view",
                "BeginningMarketValue": float(group["BeginningMarketValue"].sum()),
                "Contributions_or_Gifts": float(group["Contributions_or_Gifts"].sum()),
                "BenefitPayments_or_Distributions": float(group["BenefitPayments_or_Distributions"].sum()),
                "AdminFees": float(group["AdminFees"].sum()),
                "InvestmentManagementFees": float(group["InvestmentManagementFees"].sum()),
                "NetCashFlow": float(group["NetCashFlow"].sum()),
                "InvestmentGainLoss": float(group["InvestmentGainLoss"].sum()),
                "EndingMarketValue": float(ending),
                "QTDReturnPct": weighted("QTDReturnPct"),
                "PolicyBenchmarkQTDReturnPct": weighted("PolicyBenchmarkQTDReturnPct"),
                "ExcessQTDReturnBps": (weighted("QTDReturnPct") - weighted("PolicyBenchmarkQTDReturnPct")) * 100,
                "FYTDReturnPct": weighted("FYTDReturnPct"),
                "PolicyBenchmarkFYTDReturnPct": weighted("PolicyBenchmarkFYTDReturnPct"),
                "ExcessFYTDReturnBps": (weighted("FYTDReturnPct") - weighted("PolicyBenchmarkFYTDReturnPct")) * 100,
                "source_record_id": f"Fund_Summary|{quarter}|All",
            }
        )
    view = pd.DataFrame(rows)
    view = view.sort_values(["FundCode", "Quarter"], key=lambda col: col.map({q: i for i, q in enumerate(PERIODS)}) if col.name == "Quarter" else col)
    qoq_amounts = []
    qoq_pcts = []
    for _, row in view.iterrows():
        fund = row["FundCode"]
        q = row["Quarter"]
        prev_idx = PERIODS.index(q) - 1
        if prev_idx < 0:
            amount = row["EndingMarketValue"] - row["BeginningMarketValue"]
            base_value = row["BeginningMarketValue"]
        else:
            prev = view[(view["FundCode"] == fund) & (view["Quarter"] == PERIODS[prev_idx])]
            base_value = float(prev["EndingMarketValue"].iloc[0]) if not prev.empty else 0.0
            amount = row["EndingMarketValue"] - base_value
        qoq_amounts.append(float(amount))
        qoq_pcts.append(float(amount / base_value * 100) if base_value else 0.0)
    view["QoQAUMChange"] = qoq_amounts
    view["QoQAUMChangePct"] = qoq_pcts
    return view


def build_allocation_view(asset_allocation: pd.DataFrame, fund_view: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, row in asset_allocation.iterrows():
        fund_total = fund_view[(fund_view["FundCode"] == row["FundCode"]) & (fund_view["Quarter"] == row["Quarter"])]["EndingMarketValue"].iloc[0]
        record = row.to_dict()
        record["DollarVariance"] = float(row["EndingMarketValue"] - fund_total * row["PolicyTargetPct"] / 100)
        rows.append(record)
    for (quarter, asset), group in asset_allocation.groupby(["Quarter", "AssetClassLevel1"], sort=False):
        fund_total = fund_view[(fund_view["FundCode"] == "All") & (fund_view["Quarter"] == quarter)]["EndingMarketValue"].iloc[0]
        mv = group["EndingMarketValue"].sum()
        weight = mv if mv else 1
        policy = float((group["PolicyTargetPct"] * group["EndingMarketValue"]).sum() / weight)
        record = {
            "Quarter": quarter,
            "QuarterEndDate": group["QuarterEndDate"].iloc[0],
            "FundCode": "All",
            "FundName": "Combined Portfolio",
            "AssetClassLevel0": group["AssetClassLevel0"].iloc[0],
            "AssetClassLevel1": asset,
            "BeginningMarketValue": float(group["BeginningMarketValue"].sum()),
            "EndingMarketValue": float(mv),
            "PctOfFundTotal": float(mv / fund_total * 100) if fund_total else 0.0,
            "PolicyTargetPct": policy,
            "VarianceToTargetPct": float(mv / fund_total * 100 - policy) if fund_total else 0.0,
            "Benchmark": group["Benchmark"].iloc[0],
            "BenchmarkDataSource": group["BenchmarkDataSource"].iloc[0],
            "QTDReturnPct": float((group["QTDReturnPct"] * group["EndingMarketValue"]).sum() / weight),
            "BenchmarkQTDReturnPct": float((group["BenchmarkQTDReturnPct"] * group["EndingMarketValue"]).sum() / weight),
            "FYTDReturnPct": float((group["FYTDReturnPct"] * group["EndingMarketValue"]).sum() / weight),
            "BenchmarkFYTDReturnPct": float((group["BenchmarkFYTDReturnPct"] * group["EndingMarketValue"]).sum() / weight),
            "source_record_id": f"Asset_Allocation|{quarter}|All|{asset}",
        }
        record["ExcessQTDReturnBps"] = (record["QTDReturnPct"] - record["BenchmarkQTDReturnPct"]) * 100
        record["ExcessFYTDReturnBps"] = (record["FYTDReturnPct"] - record["BenchmarkFYTDReturnPct"]) * 100
        record["DollarVariance"] = float(mv - fund_total * policy / 100)
        rows.append(record)
    return pd.DataFrame(rows)


def build_manager_view(manager_detail: pd.DataFrame, asset_allocation: pd.DataFrame) -> pd.DataFrame:
    benchmarks = asset_allocation.set_index(["FundCode", "Quarter", "AssetClassLevel1"])
    rows = []
    for _, row in manager_detail.iterrows():
        record = row.to_dict()
        try:
            alloc = benchmarks.loc[(row["FundCode"], row["Quarter"], row["AssetClassLevel1"])]
            record["BenchmarkReturnPct"] = float(alloc["BenchmarkFYTDReturnPct"])
            record["BenchmarkQTDReturnPct"] = float(alloc["BenchmarkQTDReturnPct"])
        except KeyError:
            record["BenchmarkReturnPct"] = None
            record["BenchmarkQTDReturnPct"] = None
        record["DisplayFYTDReturnPct"] = float(row["FYTDReturnPct"])
        record["DisplayQTDReturnPct"] = float(row["QTDReturnPct"])
        record["ExcessFYTDReturnPp"] = float(row["FYTDReturnPct"] - record["BenchmarkReturnPct"]) if record["BenchmarkReturnPct"] is not None else None
        record["ExcessQTDReturnPp"] = float(row["QTDReturnPct"] - record["BenchmarkQTDReturnPct"]) if record["BenchmarkQTDReturnPct"] is not None else None
        record["DisplayReturnPct"] = record["DisplayFYTDReturnPct"]
        record["ExcessReturnPp"] = record["ExcessFYTDReturnPp"]
        ahead = 0
        for quarter in PERIODS:
            qrow = manager_detail[
                (manager_detail["FundCode"] == row["FundCode"])
                & (manager_detail["ManagerName"] == row["ManagerName"])
                & (manager_detail["AssetClassLevel1"] == row["AssetClassLevel1"])
                & (manager_detail["Quarter"] == quarter)
            ]
            if qrow.empty:
                continue
            try:
                qalloc = benchmarks.loc[(row["FundCode"], quarter, row["AssetClassLevel1"])]
            except KeyError:
                continue
            if float(qrow["QTDReturnPct"].iloc[0]) > float(qalloc["BenchmarkQTDReturnPct"]):
                ahead += 1
        record["QuartersAhead"] = ahead
        rows.append(record)
    return pd.DataFrame(rows)


def validate(canonical: dict[str, pd.DataFrame]) -> list[dict[str, Any]]:
    fund = canonical["Fund_Summary"]
    alloc = canonical["Asset_Allocation"]
    manager = canonical["Manager_Detail"]
    bench = canonical["Benchmarks_Reference"]
    validations: list[dict[str, Any]] = []
    for _, row in fund.iterrows():
        calc_net = row["Contributions_or_Gifts"] + row["BenefitPayments_or_Distributions"] + row["AdminFees"] + row["InvestmentManagementFees"]
        calc_end = row["BeginningMarketValue"] + row["NetCashFlow"] + row["InvestmentGainLoss"]
        variance = round(float(calc_end - row["EndingMarketValue"]), 6)
        net_variance = round(float(calc_net - row["NetCashFlow"]), 6)
        validations.append(
            {
                "type": "fund_roll_forward",
                "fund": row["FundCode"],
                "period": row["Quarter"],
                "variance": variance,
                "net_cash_flow_variance": net_variance,
                "tolerance": ROUNDING_TOLERANCE,
                "status": "pass" if abs(variance) <= ROUNDING_TOLERANCE and abs(net_variance) <= ROUNDING_TOLERANCE else "fail",
            }
        )
    for (fund_code, quarter), group in alloc.groupby(["FundCode", "Quarter"], sort=False):
        fund_row = fund[(fund["FundCode"] == fund_code) & (fund["Quarter"] == quarter)]
        mv_variance = round(float(group["EndingMarketValue"].sum() - fund_row["EndingMarketValue"].iloc[0]), 6) if not fund_row.empty else None
        allocation_total = round(float(group["PctOfFundTotal"].sum()), 6)
        validations.append(
            {
                "type": "allocation_total",
                "fund": fund_code,
                "period": quarter,
                "allocation_total": allocation_total,
                "market_value_variance": mv_variance,
                "tolerance": ROUNDING_TOLERANCE,
                "status": "pass"
                if abs(allocation_total - 100) <= ROUNDING_TOLERANCE and (mv_variance is None or abs(mv_variance) <= ROUNDING_TOLERANCE)
                else "fail",
            }
        )
    for (fund_code, quarter, asset), group in manager.groupby(["FundCode", "Quarter", "AssetClassLevel1"], sort=False):
        asset_row = alloc[(alloc["FundCode"] == fund_code) & (alloc["Quarter"] == quarter) & (alloc["AssetClassLevel1"] == asset)]
        variance = round(float(group["MarketValue"].sum() - asset_row["EndingMarketValue"].iloc[0]), 6) if not asset_row.empty else None
        validations.append(
            {
                "type": "manager_rollup",
                "fund": fund_code,
                "period": quarter,
                "asset_class": asset,
                "variance": variance,
                "tolerance": ROUNDING_TOLERANCE,
                "status": "pass" if variance is not None and abs(variance) <= ROUNDING_TOLERANCE else "fail",
            }
        )
    benchmark_assets = set(bench["AssetClassLevel1"])
    for asset in sorted(alloc["AssetClassLevel1"].unique()):
        validations.append({"type": "benchmark_mapping", "asset_class": asset, "status": "pass" if asset in benchmark_assets else "fail"})
    for sheet, df in canonical.items():
        key_cols = [col for col in ["Quarter", "QuarterEndDate", "FundCode", "AssetClassLevel1", "ManagerName", "FlowType", "Benchmark"] if col in df.columns]
        duplicate_count = int(df.duplicated(subset=key_cols).sum()) if key_cols else 0
        null_count = int(df.drop(columns=["_provenance"], errors="ignore").isna().sum().sum())
        validations.append({"type": "duplicates_nulls", "sheet": sheet, "duplicate_count": duplicate_count, "null_count": null_count, "status": "pass" if duplicate_count == 0 and null_count == 0 else "warn"})
    for fund_code in sorted(fund["FundCode"].unique()):
        ordered = fund[fund["FundCode"] == fund_code].sort_values("Quarter")
        prev_end = None
        for _, row in ordered.iterrows():
            if prev_end is not None:
                variance = round(float(row["BeginningMarketValue"] - prev_end), 6)
                validations.append(
                    {
                        "type": "cross_quarter_continuity",
                        "fund": fund_code,
                        "period": row["Quarter"],
                        "variance": variance,
                        "tolerance": ROUNDING_TOLERANCE,
                        "status": "pass" if abs(variance) <= ROUNDING_TOLERANCE else "fail",
                    }
                )
            prev_end = row["EndingMarketValue"]
    return validations


def _duckdb_serializable(df: pd.DataFrame) -> pd.DataFrame:
    serializable = df.drop(columns=["_provenance", "source_fields"], errors="ignore").copy()
    for column in serializable.columns:
        if serializable[column].map(lambda value: isinstance(value, (list, dict))).any():
            serializable[column] = serializable[column].map(lambda value: json.dumps(value) if isinstance(value, (list, dict)) else value)
    return serializable


def write_duckdb(
    path: Path,
    canonical: dict[str, pd.DataFrame],
    analytics: dict[str, pd.DataFrame],
    domain: dict[str, list[dict[str, Any]]] | None = None,
    metrics: dict[str, list[dict[str, Any]]] | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        path.unlink()
    con = duckdb.connect(str(path))
    try:
        for name, df in {**canonical, **analytics}.items():
            table = name.lower()
            serializable = _duckdb_serializable(df)
            con.register("df_view", serializable)
            con.execute(f"CREATE TABLE {table} AS SELECT * FROM df_view")
            con.unregister("df_view")
        for name, rows in (domain or {}).items():
            table = f"canonical_{name}"
            serializable = _duckdb_serializable(pd.DataFrame(rows))
            con.register("df_view", serializable)
            con.execute(f"CREATE TABLE {table} AS SELECT * FROM df_view")
            con.unregister("df_view")
        for name, rows in (metrics or {}).items():
            table = name
            serializable = _duckdb_serializable(pd.DataFrame(rows))
            con.register("df_view", serializable)
            con.execute(f"CREATE TABLE {table} AS SELECT * FROM df_view")
            con.unregister("df_view")
    finally:
        con.close()


def build_canonical_domain(canonical: dict[str, pd.DataFrame], analytics: dict[str, pd.DataFrame]) -> dict[str, list[dict[str, Any]]]:
    fund_summary = canonical["Fund_Summary"]
    asset_allocation = canonical["Asset_Allocation"]
    manager_detail = canonical["Manager_Detail"]
    cash_flow_detail = canonical["Cash_Flow_Detail"]
    benchmarks = canonical["Benchmarks_Reference"]
    fund_view = analytics["fund_summary_view"]
    allocation_view = analytics["asset_allocation_view"]
    manager_view = analytics["manager_performance_view"]

    funds = []
    for _, row in fund_summary[["FundCode", "FundName", "FundType"]].drop_duplicates().sort_values("FundCode").iterrows():
        fund_id = row["FundCode"]
        funds.append(
            {
                "record_id": f"FUND_{fund_id}",
                "fund_id": fund_id,
                "fund_name": row["FundName"],
                "fund_type": row["FundType"],
                "fiscal_year": "FY2026",
            }
        )

    reporting_periods = []
    period_ends = fund_summary[["Quarter", "QuarterEndDate"]].drop_duplicates().sort_values("Quarter")
    for _, row in period_ends.iterrows():
        reporting_periods.append(
            {
                "record_id": f"PERIOD_FY2026_{row['Quarter']}",
                "fiscal_year": "FY2026",
                "period_id": row["Quarter"],
                "quarter": row["Quarter"],
                "period_label": f"{row['Quarter']} FY2026",
                "period_end_date": row["QuarterEndDate"],
                "period_type": "quarter",
                "contains_quarters": [row["Quarter"]],
            }
        )
    reporting_periods.extend(
        [
            {
                "record_id": "PERIOD_FY2026_H2",
                "fiscal_year": "FY2026",
                "period_id": "H2 FY2026",
                "quarter": None,
                "period_label": "H2 FY2026 / last 6 months",
                "period_end_date": "2026-06-30",
                "period_type": "derived_horizon",
                "contains_quarters": ["Q3", "Q4"],
                "calculation_note": "Additive activity is summed across Q3 and Q4; returns are linked from QTD returns.",
            },
            {
                "record_id": "PERIOD_FY2026_FULL_YEAR",
                "fiscal_year": "FY2026",
                "period_id": "FY2026",
                "quarter": None,
                "period_label": "FY2026 full year",
                "period_end_date": "2026-06-30",
                "period_type": "fiscal_year",
                "contains_quarters": ["Q1", "Q2", "Q3", "Q4"],
            },
        ]
    )

    fund_performance = []
    for _, row in fund_view.iterrows():
        fund_id = row["FundCode"]
        quarter = row["Quarter"]
        record_id = f"FUND_PERF_FY2026_{fund_id}_{quarter}"
        fund_performance.append(
            {
                **_base_lineage(row, record_id),
                "fund_id": fund_id,
                "fund_name": row.get("FundName"),
                "period_type": "quarter_end_snapshot",
                "value_basis": "USD_millions",
                "return_basis": "percent",
                "beginning_aum": _json_value(row.get("BeginningMarketValue")),
                "ending_aum": _json_value(row.get("EndingMarketValue")),
                "qoq_aum_change": _json_value(row.get("QoQAUMChange")),
                "qoq_aum_change_pct": _json_value(row.get("QoQAUMChangePct")),
                "contributions_or_gifts": _json_value(row.get("Contributions_or_Gifts")),
                "benefit_payments_or_distributions": _json_value(row.get("BenefitPayments_or_Distributions")),
                "admin_fees": _json_value(row.get("AdminFees")),
                "investment_management_fees": _json_value(row.get("InvestmentManagementFees")),
                "net_cash_flow": _json_value(row.get("NetCashFlow")),
                "investment_gain_loss": _json_value(row.get("InvestmentGainLoss")),
                "fund_return_pct": _json_value(row.get("FYTDReturnPct")),
                "policy_benchmark_return_pct": _json_value(row.get("PolicyBenchmarkFYTDReturnPct")),
                "excess_return_pp": _json_value(float(row.get("FYTDReturnPct", 0)) - float(row.get("PolicyBenchmarkFYTDReturnPct", 0))),
                "quarter_return_pct": _json_value(row.get("QTDReturnPct")),
                "quarter_policy_benchmark_return_pct": _json_value(row.get("PolicyBenchmarkQTDReturnPct")),
                "quarter_excess_return_pp": _json_value(float(row.get("QTDReturnPct", 0)) - float(row.get("PolicyBenchmarkQTDReturnPct", 0))),
                "fytd_cumulative": True,
                "quarter_only_source_supported": True,
                "source_fields": _source_fields(
                    row,
                    [
                        "BeginningMarketValue",
                        "EndingMarketValue",
                        "QTDReturnPct",
                        "PolicyBenchmarkQTDReturnPct",
                        "FYTDReturnPct",
                        "PolicyBenchmarkFYTDReturnPct",
                    ],
                ),
            }
        )

    asset_allocations = []
    for _, row in allocation_view.iterrows():
        fund_id = row["FundCode"]
        asset_class = row["AssetClassLevel1"]
        quarter = row["Quarter"]
        record_id = f"ASSET_ALLOC_FY2026_{fund_id}_{quarter}_{str(asset_class).upper().replace(' ', '_').replace('(', '').replace(')', '')}"
        asset_allocations.append(
            {
                **_base_lineage(row, record_id),
                "fund_id": fund_id,
                "fund_name": row.get("FundName"),
                "asset_class_group": row.get("AssetClassLevel0"),
                "asset_class": asset_class,
                "benchmark_name": row.get("Benchmark"),
                "benchmark_source": row.get("BenchmarkDataSource"),
                "period_type": "quarter_end_snapshot",
                "value_basis": "USD_millions",
                "return_basis": "percent",
                "beginning_market_value": _json_value(row.get("BeginningMarketValue")),
                "ending_market_value": _json_value(row.get("EndingMarketValue")),
                "actual_allocation_pct": _json_value(row.get("PctOfFundTotal")),
                "policy_target_pct": _json_value(row.get("PolicyTargetPct")),
                "allocation_drift_pp": _json_value(row.get("VarianceToTargetPct")),
                "dollar_variance": _json_value(row.get("DollarVariance")),
                "asset_return_pct": _json_value(row.get("FYTDReturnPct")),
                "asset_benchmark_return_pct": _json_value(row.get("BenchmarkFYTDReturnPct")),
                "asset_excess_return_pp": _json_value(float(row.get("FYTDReturnPct", 0)) - float(row.get("BenchmarkFYTDReturnPct", 0))),
                "quarter_asset_return_pct": _json_value(row.get("QTDReturnPct")),
                "quarter_asset_benchmark_return_pct": _json_value(row.get("BenchmarkQTDReturnPct")),
                "quarter_asset_excess_return_pp": _json_value(float(row.get("QTDReturnPct", 0)) - float(row.get("BenchmarkQTDReturnPct", 0))),
                "snapshot_measure": True,
                "fytd_cumulative_return": True,
                "quarter_only_return_source_supported": True,
                "source_fields": _source_fields(row, ["AssetClassLevel0", "AssetClassLevel1", "PctOfFundTotal", "PolicyTargetPct", "VarianceToTargetPct"]),
            }
        )

    manager_names = manager_detail[["ManagerName", "AssetClassLevel1", "VehicleType", "InceptionDate"]].drop_duplicates().sort_values("ManagerName")
    managers = []
    for _, row in manager_names.iterrows():
        manager_id = str(row["ManagerName"]).upper().replace(" ", "_").replace("-", "_")
        managers.append(
            {
                "record_id": f"MANAGER_{manager_id}",
                "manager_id": manager_id,
                "manager_name": row["ManagerName"],
                "asset_class": row["AssetClassLevel1"],
                "vehicle_type": row["VehicleType"],
                "inception_date": row["InceptionDate"],
                "fiscal_year": "FY2026",
            }
        )

    manager_performance = []
    for _, row in manager_view.iterrows():
        manager_id = str(row["ManagerName"]).upper().replace(" ", "_").replace("-", "_")
        fund_id = row["FundCode"]
        quarter = row["Quarter"]
        record_id = f"MANAGER_PERF_FY2026_{fund_id}_{quarter}_{manager_id}"
        manager_performance.append(
            {
                **_base_lineage(row, record_id),
                "fund_id": fund_id,
                "fund_name": row.get("FundName"),
                "manager_id": manager_id,
                "manager_name": row.get("ManagerName"),
                "asset_class": row.get("AssetClassLevel1"),
                "vehicle_type": row.get("VehicleType"),
                "inception_date": row.get("InceptionDate"),
                "period_type": "quarter_end_snapshot",
                "value_basis": "USD_millions",
                "return_basis": "percent",
                "manager_aum": _json_value(row.get("MarketValue")),
                "pct_of_asset_class": _json_value(row.get("PctOfAssetClass")),
                "manager_return_pct": _json_value(row.get("FYTDReturnPct")),
                "manager_benchmark_return_pct": _json_value(row.get("BenchmarkReturnPct")),
                "manager_excess_return_pp": _json_value(row.get("ExcessFYTDReturnPp")),
                "quarter_manager_return_pct": _json_value(row.get("QTDReturnPct")),
                "quarter_manager_benchmark_return_pct": _json_value(row.get("BenchmarkQTDReturnPct")),
                "quarter_manager_excess_return_pp": _json_value(row.get("ExcessQTDReturnPp")),
                "quarters_outperforming": _json_value(row.get("QuartersAhead")),
                "fytd_cumulative_return": True,
                "quarter_only_return_source_supported": True,
                "source_fields": _source_fields(row, ["ManagerName", "MarketValue", "PctOfAssetClass", "QTDReturnPct", "FYTDReturnPct"]),
            }
        )

    cash_flows = []
    for _, row in cash_flow_detail.iterrows():
        fund_id = row["FundCode"]
        quarter = row["Quarter"]
        flow_type = row["FlowType"]
        record_id = f"CASH_FLOW_FY2026_{fund_id}_{quarter}_{str(flow_type).upper()}"
        amount = float(row["Amount"])
        cash_flows.append(
            {
                **_base_lineage(row, record_id),
                "fund_id": fund_id,
                "fund_name": row.get("FundName"),
                "flow_type": flow_type,
                "amount": amount,
                "cash_flow_direction": "inflow" if amount > 0 else "outflow" if amount < 0 else "flat",
                "period_type": "quarter_activity",
                "value_basis": "USD_millions",
                "cumulative_fytd": False,
                "quarter_only_source_supported": True,
                "source_fields": _source_fields(row, ["FlowType", "Amount"]),
            }
        )

    benchmark_records = []
    for _, row in benchmarks.iterrows():
        asset_class = row["AssetClassLevel1"]
        record_id = f"BENCHMARK_{str(asset_class).upper().replace(' ', '_').replace('(', '').replace(')', '')}"
        benchmark_records.append(
            {
                **_base_lineage(row, record_id),
                "asset_class_group": row.get("AssetClassLevel0"),
                "asset_class": asset_class,
                "benchmark_name": row.get("Benchmark"),
                "benchmark_source": row.get("DataSource"),
                "fiscal_year": "FY2026",
                "source_fields": _source_fields(row, ["AssetClassLevel0", "AssetClassLevel1", "Benchmark", "DataSource"]),
            }
        )

    return {
        "funds": funds,
        "reporting_periods": reporting_periods,
        "fund_performance": fund_performance,
        "asset_allocations": asset_allocations,
        "managers": managers,
        "manager_performance": manager_performance,
        "cash_flows": cash_flows,
        "benchmarks": benchmark_records,
    }


def build_model(data_dir: Path, output_dir: Path, store_path: Path) -> dict[str, Any]:
    canonical, books, duplicate_records, raw_samples = load_sources(data_dir)
    fund_view = build_fund_view(canonical["Fund_Summary"])
    allocation_view = build_allocation_view(canonical["Asset_Allocation"], fund_view)
    manager_view = build_manager_view(canonical["Manager_Detail"], canonical["Asset_Allocation"])
    validations = validate(canonical)
    asset_classes = sorted(canonical["Asset_Allocation"]["AssetClassLevel1"].unique().tolist())
    funds = canonical["Fund_Summary"][["FundCode", "FundName", "FundType"]].drop_duplicates().sort_values("FundCode")
    managers = sorted(canonical["Manager_Detail"]["ManagerName"].unique().tolist())
    benchmark_coverage = [
        {"asset_class": asset, "has_benchmark": bool((canonical["Benchmarks_Reference"]["AssetClassLevel1"] == asset).any())}
        for asset in asset_classes
    ]
    analytics = {
        "fund_summary_view": fund_view,
        "asset_allocation_view": allocation_view,
        "manager_performance_view": manager_view,
    }
    for horizon in HORIZONS:
        slug = horizon.lower().replace(" ", "_")
        analytics[f"fund_horizon_{slug}"] = fund_horizon_view(fund_view, horizon)
        analytics[f"allocation_horizon_{slug}"] = allocation_horizon_view(allocation_view, horizon)
        analytics[f"manager_horizon_{slug}"] = manager_horizon_view(manager_view, horizon)
    comparisons = {
        "q4_vs_q3": q4_vs_q3(allocation_view, manager_view),
        "h2_vs_h1": h2_vs_h1(fund_view, manager_view),
        "q1_to_q4_trajectory": q1_to_q4_trajectory(allocation_view, manager_view),
    }
    research_by_horizon = build_research_by_horizon(canonical, analytics)
    canonical_domain = build_canonical_domain(canonical, analytics)
    metric_registry = metric_registry_records()
    metric_values = build_metric_values(canonical, analytics, validations)
    semantic_layer = build_semantic_layer(asset_classes, managers, metric_registry)
    write_duckdb(
        store_path,
        canonical,
        analytics,
        canonical_domain,
        {"metric_registry": metric_registry, "metric_values": metric_values},
    )
    model = {
        "generated_at": pd.Timestamp.now().isoformat(timespec="seconds"),
        "methodology": {
            "units": "USD millions ($M); returns in percent; excess returns in basis points in source and percentage points in UI.",
            "fytd": "Each workbook is a fiscal-year-to-date snapshot through its file quarter; canonical quarter records are deduplicated by quarter-end date.",
            "roll_forward": "Ending Market Value = Beginning Market Value + Net Cash Flow + Investment Gain/Loss.",
            "net_cash_flow": "Contributions/Gifts + Benefit Payments/Distributions + Admin Fees + Investment Management Fees.",
            "policy_benchmark": "Fund policy benchmark return is target-weighted asset-class benchmark return.",
            "quarter_only_activity": "Quarter-only activity is read from quarter rows or derived only for additive activity/snapshot deltas. Quarterly returns are not derived by subtracting FYTD returns.",
        },
        "files": books,
        "dimensions": {
            "funds": _records(funds),
            "periods": ["Q1", "Q2", "Q3", "Q4", "FY2026"],
            "metric_periods": METRIC_PERIODS,
            "asset_classes": asset_classes,
            "managers": managers,
            "benchmark_coverage": benchmark_coverage,
        },
        "records": {
            "fund_summary": _records(canonical["Fund_Summary"]),
            "asset_allocation": _records(canonical["Asset_Allocation"]),
            "manager_detail": _records(canonical["Manager_Detail"]),
            "cash_flow_detail": _records(canonical["Cash_Flow_Detail"]),
            "benchmarks_reference": _records(canonical["Benchmarks_Reference"]),
            "raw_export_sample": raw_samples,
        },
        "analytics": {name: _records(frame) for name, frame in analytics.items()},
        "canonical": canonical_domain,
        "metric_registry": metric_registry,
        "metric_values": metric_values,
        "semantic_layer": semantic_layer,
        "business_tools": {"schemas": tool_schemas()},
        "canonical_schema": {
            "funds": ["record_id", "fund_id", "fund_name", "fund_type", "fiscal_year"],
            "reporting_periods": ["record_id", "fiscal_year", "period_id", "quarter", "period_label", "period_end_date", "period_type", "contains_quarters"],
            "fund_performance": [
                "record_id",
                "fund_id",
                "fiscal_year",
                "quarter",
                "ending_aum",
                "fund_return_pct",
                "policy_benchmark_return_pct",
                "excess_return_pp",
                "net_cash_flow",
                "investment_gain_loss",
            ],
            "asset_allocations": [
                "record_id",
                "fund_id",
                "quarter",
                "asset_class",
                "ending_market_value",
                "actual_allocation_pct",
                "policy_target_pct",
                "allocation_drift_pp",
                "dollar_variance",
            ],
            "managers": ["record_id", "manager_id", "manager_name", "asset_class", "vehicle_type", "inception_date"],
            "manager_performance": [
                "record_id",
                "fund_id",
                "quarter",
                "manager_id",
                "manager_name",
                "asset_class",
                "manager_aum",
                "manager_return_pct",
                "manager_benchmark_return_pct",
                "manager_excess_return_pp",
            ],
            "cash_flows": ["record_id", "fund_id", "quarter", "flow_type", "amount", "cash_flow_direction"],
            "benchmarks": ["record_id", "asset_class", "benchmark_name", "benchmark_source"],
        },
        "research": {
            "attribution_status": "Attribution not fully supported. The dataset supports relative performance drivers, but not formal holdings-level or average-weight attribution.",
            "horizons": research_by_horizon,
            "comparisons": comparisons,
            "summary": research_by_horizon["FY2026"]["summary"],
            "candidates": research_by_horizon["FY2026"]["candidates"],
            "final_signals": research_by_horizon["FY2026"]["final_signals"],
        },
        "audit": {
            "duplicate_records": duplicate_records,
            "validations": validations,
            "duckdb_store": str(store_path),
            "raw_export_assessment": "RAW_Export_Extract contains source-system metadata blocks, JSON config, blank spacer rows, cryptic platform column IDs and display-label rows before data. The structured sheets are safer for the primary reporting pipeline; raw parsing would require skipping metadata, selecting the technical header row, coercing types, and mapping platform IDs to reporting fields.",
        },
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    json_text = json.dumps(model, indent=2, ensure_ascii=False)
    (output_dir / "beacon-data.json").write_text(json_text, encoding="utf-8")
    (output_dir / "beacon-data.js").write_text(f"window.BEACON_DATA = {json_text};\n", encoding="utf-8")
    return model


def main() -> None:
    parser = argparse.ArgumentParser(description="Build Beacon normalized reporting data.")
    parser.add_argument("--data-dir", type=Path, default=Path("Data"))
    parser.add_argument("--output-dir", type=Path, default=Path("public/data"))
    parser.add_argument("--store-path", type=Path, default=Path("public/data/beacon.duckdb"))
    args = parser.parse_args()
    model = build_model(args.data_dir, args.output_dir, args.store_path)
    failures = [v for v in model["audit"]["validations"] if v["status"] == "fail"]
    print(
        json.dumps(
            {
                "fund_rows": len(model["records"]["fund_summary"]),
                "allocation_rows": len(model["records"]["asset_allocation"]),
                "manager_rows": len(model["records"]["manager_detail"]),
                "cash_flow_rows": len(model["records"]["cash_flow_detail"]),
                "benchmark_rows": len(model["records"]["benchmarks_reference"]),
                "validation_failures": len(failures),
            },
            indent=2,
        )
    )
